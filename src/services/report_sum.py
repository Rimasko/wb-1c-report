from collections import defaultdict
from pathlib import Path
from pprint import pprint
from typing import Final

import active as active
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from settings import Settings


class ReportSumService:
    class ReportException(Exception):
        ...

    EXCEL_EXTENSIONS: Final[list[str]] = ['.xlsx', '.xls']

    def __init__(self):
        self.settings = Settings()

    def _get_titles(self, active_sheet: Worksheet) -> dict[str, str]:
        return {
            col: active_sheet[f'{col}{self.settings.title_row_index}'].value
            for col in (self.settings.unique_cols + self.settings.sum_cols)
        }

    def _report_processing(self, active_sheet: Worksheet, table_titles: dict) -> dict[str, ]:
        table = dict()

        for i in range(self.settings.values_row_start_index, active_sheet.max_row):
            unique_key = ':'.join(active_sheet[f'{col}{i}'].value for col in self.settings.unique_cols)

            if unique_key not in table:
                table[unique_key] = {
                    **{table_titles[col]: active_sheet[f'{col}{i}'].value for col in self.settings.unique_cols},
                    **{table_titles[col]: active_sheet[f'{col}{i}'].value for col in self.settings.sum_cols},
                }
            else:
                row = {table_titles[col]: active_sheet[f'{col}{i}'].value for col in self.settings.sum_cols}

                for col_name, col_value in row.items():
                    table[unique_key][col_name] += col_value

        return dict(sorted(table.items(), key=lambda item: list(item[1].values())[2], reverse=True))

    def get_report_workbook(self, report_map: dict, table_titles: dict) -> Workbook:
        wb = Workbook()
        sheet = wb.active
        title_column_letter = {}
        column_lengths = {}

        sheet.column_dimensions['A'].bestFit = True

        for i, value in enumerate(table_titles.values()):
            sheet[f'{chr(ord("A") + i)}{self.settings.title_row_index}'] = value
            title_column_letter[value] = chr(ord("A") + i)
            column_lengths[chr(ord("A") + i)] = len(value)

        for i, row in enumerate(report_map.values(), self.settings.values_row_start_index):
            for column_name, value in row.items():
                letter = title_column_letter[column_name]
                sheet[f'{letter}{i}'] = value
                column_lengths[letter] = max(column_lengths[letter], len(str(value)))

        for letter, length in column_lengths.items():
            sheet.column_dimensions[letter].width = length

        return wb

    def get_sum_report(self, wildberries_report_file: Path, save_file_path: Path):
        if (
            not wildberries_report_file.is_file() or
            wildberries_report_file.suffix not in self.EXCEL_EXTENSIONS
        ):
            raise self.ReportException('NOT excel file')

        report_workbook = load_workbook(filename=wildberries_report_file)
        active_sheet = report_workbook.active
        table_titles = self._get_titles(active_sheet)
        report_map = self._report_processing(active_sheet, table_titles)

        wb = self.get_report_workbook(report_map, table_titles)
        wb.save(save_file_path)
